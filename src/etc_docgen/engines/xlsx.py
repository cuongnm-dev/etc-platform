#!/usr/bin/env python3
"""
fill_xlsx_engine.py — Template-schema-driven Excel fill engine.

Reads:
  - template (.xlsx)        — ETC-provided, immutable
  - schema  (.yaml)         — one-time analysis of template structure
  - content-data (.json)    — data produced by tdoc-data-writer

Writes:
  - output (.xlsx) — template copy with cells filled per schema

Guarantees:
  - NEVER writes to a formula cell declared in schema.preserve.formula_cells
  - NEVER writes to a MergedCell that is not the top-left anchor
  - Clears data_table.clear_columns before writing new rows
  - Runs validators after write; aborts (non-zero exit) if any fail

Usage:
  python fill_xlsx_engine.py \
    --template   test-case.xlsx \
    --schema     schemas/test-case.xlsx.schema.yaml \
    --data       content-data.json \
    --output     out/kich-ban-kiem-thu.xlsx \
    [--report    out/xlsx-fill-report.json]
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import tempfile
import unicodedata
import zipfile
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

try:
    import yaml
except ImportError:
    print("ERROR: PyYAML required. Run: pip install pyyaml", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ─────────────────────────── Fill report ───────────────────────────


@dataclass
class FillReport:
    writes_done: int = 0
    writes_skipped_formula: int = 0
    writes_skipped_protected: int = 0
    rows_cleared: int = 0
    rows_written: int = 0
    warnings: list[str] = field(default_factory=list)
    validator_failures: list[str] = field(default_factory=list)
    per_sheet: dict[str, dict[str, int]] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "writes_done": self.writes_done,
            "writes_skipped_formula": self.writes_skipped_formula,
            "writes_skipped_protected": self.writes_skipped_protected,
            "rows_cleared": self.rows_cleared,
            "rows_written": self.rows_written,
            "warnings": self.warnings,
            "validator_failures": self.validator_failures,
            "per_sheet": self.per_sheet,
            "status": "ok" if not self.validator_failures else "validation_failed",
        }


# ─────────────────────────── JSONPath lite ───────────────────────────

_DOT_RE = re.compile(r"\.(?=[^.\[]|$)")


def resolve(data: dict, path: str | None) -> Any:
    """Resolve a '$.a.b.c' path in `data`. Returns None if any segment missing.

    Supports dots only (no wildcards, no [index] — keep it simple & explicit).
    """
    if path is None:
        return None
    if not path.startswith("$"):
        return path  # literal
    segments = [s for s in path.lstrip("$").lstrip(".").split(".") if s]
    cur: Any = data
    for seg in segments:
        if isinstance(cur, dict) and seg in cur:
            cur = cur[seg]
        else:
            return None
    return cur


# ─────────────────────────── Transforms ───────────────────────────


def apply_transform(value: Any, transform: str, schema: dict) -> Any:
    """Apply a named transform. Returns transformed value or original if unknown."""
    if value is None:
        return None

    if transform == "upper":
        return str(value).upper()
    if transform == "lower":
        return str(value).lower()

    m = re.match(r"truncate\((\d+)\)", transform)
    if m:
        n = int(m.group(1))
        s = str(value)
        return s if len(s) <= n else s[: n - 3] + "..."

    m = re.match(r"join\(['\"](.+?)['\"]\)", transform)
    if m:
        sep = m.group(1)
        if isinstance(value, list):
            return sep.join(str(x) for x in value)
        return str(value)

    if transform == "numbered_join":
        # value expected: list[str] or list[dict{no, text|action|expected}]
        if isinstance(value, list):
            lines = []
            for i, item in enumerate(value, 1):
                if isinstance(item, dict):
                    no = item.get("no", i)
                    txt = item.get("text") or item.get("action") or item.get("expected") or ""
                    lines.append(f"Bước {no}: {txt}")
                else:
                    lines.append(f"Bước {i}: {item}")
            return "\n".join(lines)
        return str(value)

    if transform == "priority_map":
        pm = schema.get("priority_map", {})
        default = schema.get("priority_default", "Normal")
        # Try direct match first, then normalize Vietnamese diacritics
        s = str(value)
        if s in pm:
            return pm[s]
        # Normalize: strip diacritics for fuzzy match
        s_norm = (
            "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
            .lower()
            .strip()
        )
        for key, mapped in pm.items():
            k_norm = (
                "".join(
                    c for c in unicodedata.normalize("NFKD", key) if not unicodedata.combining(c)
                )
                .lower()
                .strip()
            )
            if s_norm == k_norm:
                return mapped
        return default

    return value  # unknown transform — pass through


def render_template_string(tpl: str, data: dict) -> str:
    """Replace {a.b.c} tokens with resolve(data, '$.a.b.c')."""

    def sub(m: re.Match) -> str:
        path = "$." + m.group(1)
        v = resolve(data, path)
        return str(v) if v is not None else ""

    return re.sub(r"\{([a-zA-Z0-9_.]+)\}", sub, tpl)


def resolve_write_value(write: dict, data: dict, schema: dict) -> Any:
    """Resolve a write entry's final value.

    Priority:
      1. Hardcoded value       (write.value)
      2. Source from data      (write.source) — PRIMARY for data-driven fields
      3. Fallback source       (write.fallback_source) — extract sub-key from alternate field
      4. Template fallback     (write.template) — used only when source missing
      5. Default               (write.default)
    """
    if "value" in write:
        return write["value"]

    # Try source first (data-driven)
    if "source" in write:
        v = resolve(data, write["source"])
        if v is not None and v != "" and v != []:
            if "transform" in write:
                v = apply_transform(v, write["transform"], schema)
            return v
        # source empty → try fallback

    # Try fallback_source — extract sub-key from an alternate list field
    # Use case: $.expected is empty, but $.steps[].expected has embedded values
    if "fallback_source" in write:
        fb = resolve(data, write["fallback_source"])
        fb_key = write.get("fallback_key", "")
        if isinstance(fb, list) and fb_key:
            extracted = [
                {"no": item.get("no", i + 1), fb_key: item[fb_key]}
                for i, item in enumerate(fb)
                if isinstance(item, dict) and item.get(fb_key)
            ]
            if extracted:
                if "transform" in write:
                    return apply_transform(extracted, write["transform"], schema)
                return extracted

    # Try template (built from data)
    if "template" in write:
        v = render_template_string(write["template"], data)
        if v:
            return v

    return write.get("default")


# ─────────────────────────── Safe cell write ───────────────────────────


def get_merge_anchor(ws, cell_ref: str) -> str:
    """If cell_ref is inside a merged range, return the top-left anchor; else return cell_ref."""
    if not isinstance(ws[cell_ref], MergedCell):
        return cell_ref
    for rng in ws.merged_cells.ranges:
        if cell_ref in rng:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return cell_ref  # shouldn't happen


def is_formula_cell(cell_ref: str, sheet_schema: dict) -> bool:
    formula_cells = sheet_schema.get("_formula_cells_cache")
    if formula_cells is None:
        preserve = sheet_schema.get("preserve", {}) or {}
        formula_cells = frozenset(preserve.get("formula_cells") or [])
        sheet_schema["_formula_cells_cache"] = formula_cells
    return cell_ref in formula_cells


def safe_write(ws, cell_ref: str, value: Any, sheet_schema: dict, report: FillReport) -> bool:
    """Write value to cell unless it's a formula cell. Routes through merge anchor."""
    if is_formula_cell(cell_ref, sheet_schema):
        report.writes_skipped_formula += 1
        report.warnings.append(
            f"Skipped formula-cell write: {ws.title}!{cell_ref} (would destroy formula)"
        )
        return False
    anchor = get_merge_anchor(ws, cell_ref)
    if is_formula_cell(anchor, sheet_schema):
        report.writes_skipped_formula += 1
        return False
    ws[anchor] = value
    report.writes_done += 1
    return True


# ─────────────────────────── Sheet fill ───────────────────────────


def process_sheet_writes(ws, sheet_schema: dict, data: dict, schema: dict, report: FillReport):
    """Process sheet.writes + sheet.clear_cells."""
    sheet_stats = report.per_sheet.setdefault(ws.title, {"writes": 0, "rows": 0})

    # Clear first
    for cell_ref in sheet_schema.get("clear_cells", []) or []:
        if is_formula_cell(cell_ref, sheet_schema):
            continue
        anchor = get_merge_anchor(ws, cell_ref)
        ws[anchor] = None

    # Writes
    for write in sheet_schema.get("writes", []) or []:
        cell_ref = write["cell"]
        value = resolve_write_value(write, data, schema)
        if value is None:
            continue
        if safe_write(ws, cell_ref, value, sheet_schema, report):
            sheet_stats["writes"] += 1


def _unmerge_data_region(ws, start_row: int, end_row: int, report: FillReport) -> int:
    """Unmerge any merged range that falls within [start_row, end_row].

    Template often has sample data with feature-group header merges (A:M).
    Those merges must be removed before we fill flat per-TC data;
    otherwise writing to cells inside the merge is silently skipped.

    Returns count of unmerged ranges.
    """
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        # overlap check
        if rng.min_row >= start_row and rng.max_row <= end_row:
            to_unmerge.append(str(rng))
    for r in to_unmerge:
        ws.unmerge_cells(r)
    if to_unmerge:
        report.warnings.append(
            f"{ws.title}: unmerged {len(to_unmerge)} ranges in data region "
            f"rows {start_row}-{end_row} (template sample-data leftovers)"
        )
    return len(to_unmerge)


# ─────────────────────────── Row formatting helpers ───────────────────────────


def _capture_row_styles(ws, row_num: int, max_col: int = 13) -> dict[int, dict]:
    """Capture cell styles from a reference row for later replication."""
    styles: dict[int, dict] = {}
    for col in range(1, max_col + 1):
        c = ws.cell(row=row_num, column=col)
        styles[col] = {
            "font": copy(c.font),
            "fill": copy(c.fill),
            "border": copy(c.border),
            "alignment": copy(c.alignment),
            "number_format": c.number_format,
            "protection": copy(c.protection),
        }
    return styles


def _apply_row_styles(ws, row_num: int, styles: dict[int, dict]) -> None:
    """Apply previously captured styles to every cell in a row."""
    for col, st in styles.items():
        c = ws.cell(row=row_num, column=col)
        c.font = st["font"]
        c.fill = st["fill"]
        c.border = st["border"]
        c.alignment = st["alignment"]
        c.number_format = st["number_format"]
        c.protection = st["protection"]


def _build_auto_id_formula(row_num: int, auto_id_cfg: dict) -> str:
    """Generate the auto-numbering formula for column A.

    Replicates the ETC template pattern:
      =IF(OR(B{r}<>"",D{r}<>""),
           "["&TEXT($B$1,"##")&"-"&TEXT(COUNTA({anchor}:D{r}),"##")&"]","")
    """
    anchor = auto_id_cfg["count_anchor"]
    return (
        f'=IF(OR(B{row_num}<>"",D{row_num}<>""),'
        f'"["&TEXT($B$1,"##")&"-"&TEXT(COUNTA({anchor}:D{row_num}),"##")&"]","")'
    )


def _write_group_header_row(
    ws,
    row_num: int,
    text: str,
    style_cfg: dict,
    max_col: int = 13,
) -> None:
    """Write a merged, styled group-header or section-header row."""
    last_col = get_column_letter(max_col)
    ws.merge_cells(f"A{row_num}:{last_col}{row_num}")

    cell = ws.cell(row=row_num, column=1)
    cell.value = text
    cell.font = Font(name="Times New Roman", size=13, bold=True)

    fill_color = style_cfg.get("fill_color", "FF8EA9DB")
    cell.fill = PatternFill(
        start_color=fill_color,
        end_color=fill_color,
        fill_type="solid",
    )

    vert = style_cfg.get("vertical", "center")
    wrap = style_cfg.get("wrap_text", False)
    cell.alignment = Alignment(vertical=vert, wrap_text=wrap)

    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_height = style_cfg.get("row_height")
    if row_height:
        ws.row_dimensions[row_num].height = row_height


# ─────────────────────────── Data table fill ───────────────────────────


def process_data_table(ws, sheet_schema: dict, data: dict, schema: dict, report: FillReport):
    """Clear data region then write rows with formatting + auto-ID.

    Supports three row types in the data source:
      - Normal test-case rows (default)
      - Feature group headers  (``_type: "feature_group"``)  — merged, bold, colored
      - Section headers        (``_type: "section_header"``)  — merged, bold, light colored

    When ``data_table.format`` is present in schema:
      - Copies cell styles from ``format.reference_row``
      - Generates column-A auto-numbering formulas (``format.auto_id``)
      - Renders group/section header rows (``format.group_styles``)
    """
    dt = sheet_schema.get("data_table")
    if not dt:
        return
    start_row = dt["start_row"]
    end_cap = dt.get("end_row_hard_cap", 600)
    clear_cols = dt.get("clear_columns", [])
    fmt = dt.get("format")

    rows_data = resolve(data, dt["source"]) or []
    if not isinstance(rows_data, list):
        report.warnings.append(
            f"{ws.title}: data_table.source '{dt['source']}' resolved to non-list; skipped"
        )
        return

    # ── Capture reference row styles before clearing ──
    ref_styles = None
    if fmt:
        ref_row = fmt.get("reference_row", start_row)
        ref_styles = _capture_row_styles(ws, ref_row, ws.max_column)

    # ── STEP 1: Unmerge template sample-data merges in data region ──
    _unmerge_data_region(ws, start_row, end_cap, report)

    # ── STEP 1b: Unhide all rows in data region ──
    for rn in list(ws.row_dimensions):
        if start_row <= rn <= end_cap:
            ws.row_dimensions[rn].hidden = False

    # ── STEP 2: Clear clear_columns in data region ──
    protected = set(sheet_schema.get("preserve", {}).get("protected_columns_in_data") or [])
    safe_clear_cols = [c for c in clear_cols if c not in protected]

    # Clear only rows we'll overwrite; STEP 4 handles template residuals beyond last data row
    data_clear_end = start_row + len(rows_data)  # exclusive upper bound
    for row_num in range(start_row, min(data_clear_end, end_cap + 1)):
        for col_letter in safe_clear_cols:
            cell_ref = f"{col_letter}{row_num}"
            if is_formula_cell(cell_ref, sheet_schema):
                continue
            if isinstance(ws[cell_ref], MergedCell):
                continue
            ws[cell_ref] = None
    report.rows_cleared += min(data_clear_end, end_cap + 1) - start_row

    # ── STEP 3: Write rows ──
    sheet_stats = report.per_sheet.setdefault(ws.title, {"writes": 0, "rows": 0})
    auto_id = fmt.get("auto_id") if fmt else None
    group_styles = fmt.get("group_styles", {}) if fmt else {}

    for i, row_item in enumerate(rows_data):
        target_row = start_row + i
        if target_row > end_cap:
            report.warnings.append(
                f"{ws.title}: row {target_row} exceeds end_row_hard_cap={end_cap}; "
                f"{len(rows_data) - i} rows skipped"
            )
            break

        row_type = row_item.get("_type") if isinstance(row_item, dict) else None

        # ── Group / section header row ──
        if row_type in ("feature_group", "section_header"):
            style_cfg = group_styles.get(row_type, {})
            _write_group_header_row(ws, target_row, row_item.get("title", ""), style_cfg)
            sheet_stats["rows"] += 1
            report.rows_written += 1
            continue

        # ── Normal data row: apply reference formatting ──
        if ref_styles:
            _apply_row_styles(ws, target_row, ref_styles)

        data_rh = fmt.get("data_row_height") if fmt else None
        if data_rh:
            ws.row_dimensions[target_row].height = data_rh

        # ── Auto-ID formula in column A ──
        if auto_id and auto_id.get("enabled"):
            formula = _build_auto_id_formula(target_row, auto_id)
            ws.cell(row=target_row, column=1).value = formula

        # ── Data columns ──
        for col_letter, col_spec in dt["row_template"].items():
            cell_ref = f"{col_letter}{target_row}"
            if col_letter in protected:
                continue
            value = resolve_write_value(col_spec, row_item, schema)
            if value is None:
                continue
            if isinstance(ws[cell_ref], MergedCell):
                report.warnings.append(
                    f"{ws.title}!{cell_ref}: still MergedCell after unmerge; skipped"
                )
                continue
            safe_write(ws, cell_ref, value, sheet_schema, report)
        sheet_stats["rows"] += 1
        report.rows_written += 1

    # ── STEP 4: Strip trailing rows (template sample-data residuals) ──
    last_written = start_row + len(rows_data) - 1 if rows_data else start_row - 1
    for row_num in range(last_written + 1, end_cap + 1):
        rd = ws.row_dimensions.get(row_num)
        has_residual = rd is not None and rd.customHeight
        if not has_residual:
            for col in range(1, 14):
                c = ws.cell(row=row_num, column=col)
                if not isinstance(c, MergedCell) and (
                    c.fill.fill_type or (c.border.bottom and c.border.bottom.style)
                ):
                    has_residual = True
                    break
        if not has_residual:
            break  # past template sample rows
        for col in range(1, 14):
            cell = ws.cell(row=row_num, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment()
        if rd:
            rd.height = None


# ─────────────────────────── Validators ───────────────────────────


def run_validators(wb, schema: dict, report: FillReport):
    for v in schema.get("validators", []) or []:
        vtype = v["type"]
        name = v.get("name", vtype)
        try:
            if vtype == "sheet_names_match":
                expected = set(v["expected"])
                actual = set(wb.sheetnames)
                missing = expected - actual
                if missing:
                    report.validator_failures.append(f"{name}: missing sheets {missing}")

            elif vtype == "formula_cells_intact":
                ws = wb[v["sheet"]]
                for cell_ref in v["cells"]:
                    val = ws[cell_ref].value
                    if not (isinstance(val, str) and val.startswith("=")):
                        report.validator_failures.append(
                            f"{name}: {v['sheet']}!{cell_ref} is not a formula (got: {val!r})"
                        )

            elif vtype == "cell_values_in_set":
                ws = wb[v["sheet"]]
                col = v["column"]
                allowed = set(v["allowed"])
                for row in range(v["start_row"], v["end_row"] + 1):
                    val = ws[f"{col}{row}"].value
                    if val not in allowed:
                        report.validator_failures.append(
                            f"{name}: {v['sheet']}!{col}{row} = {val!r} not in allowed set"
                        )
                        break  # only report first occurrence

            elif vtype == "cells_not_equal":
                ws = wb[v["sheet"]]
                for chk in v["checks"]:
                    val = ws[chk["cell"]].value
                    if val == chk["must_not_equal"]:
                        report.validator_failures.append(
                            f"{name}: {v['sheet']}!{chk['cell']} still equals template placeholder"
                        )

            elif vtype == "row_count_min":
                ws = wb[v["sheet"]]
                count = 0
                for row in ws.iter_rows(
                    min_row=v["start_row"],
                    max_row=v["start_row"] + 1000,
                    min_col=column_index_from_string(v["column"]),
                    max_col=column_index_from_string(v["column"]),
                ):
                    if row[0].value:
                        count += 1
                if count < v["min_count"]:
                    # Soft: if no rows at all, it's a warning (possibly no test_cases
                    # data provided). Hard fail only when partial rows present but below min.
                    msg = (
                        f"{name}: {v['sheet']} col {v['column']} has {count} filled rows, "
                        f"expected >= {v['min_count']}"
                    )
                    if count == 0 and v.get("soft_if_empty", True):
                        report.warnings.append(msg + " (no data provided — soft warning)")
                    else:
                        report.validator_failures.append(msg)
            else:
                report.warnings.append(f"Unknown validator type: {vtype}")
        except Exception as e:
            report.validator_failures.append(f"{name}: validator crashed: {e}")


# ─────────────────── Data Validation expansion ───────────────────


def _expand_data_validations(ws, sheet_schema: dict, data: dict, report: FillReport):
    """Expand standard data-validation sqrefs to cover all written data rows.

    Template validations only cover the sample rows (e.g. I11, I15:I18).
    After we write N actual rows, we must extend those ranges to cover
    start_row..last_data_row for each relevant column.
    """
    dt = sheet_schema.get("data_table")
    if not dt:
        return
    dvs = ws.data_validations
    if not dvs or not dvs.dataValidation:
        return

    start_row = dt["start_row"]
    rows_data = resolve(data, dt["source"]) or []
    if not rows_data:
        return
    last_row = start_row + len(rows_data) - 1

    for dv in dvs.dataValidation:
        if dv.formula1 is None and dv.type is None:
            continue  # skip empty validations
        old_sqref = str(dv.sqref)
        new_ranges = []
        seen_cols: set[str] = set()
        changed = False

        for part in old_sqref.split():
            # Parse range like "I15:I18" or "I11" or "K15:K18"
            m = re.match(r"^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", part)
            if not m:
                new_ranges.append(part)
                continue

            col_start = m.group(1)
            row_start_ref = int(m.group(2))
            col_end = m.group(3) or col_start
            row_end_ref = int(m.group(4)) if m.group(4) else row_start_ref

            # Only expand ranges that overlap with the data region
            if row_start_ref < start_row or col_start != col_end:
                new_ranges.append(part)
                continue

            # Deduplicate: one contiguous range per column pair
            col_key = f"{col_start}:{col_end}"
            if col_key in seen_cols:
                changed = True
                continue
            seen_cols.add(col_key)

            # Expand to cover all data rows
            new_range = f"{col_start}{start_row}:{col_end}{last_row}"
            if new_range != part:
                changed = True
            new_ranges.append(new_range)

        if changed:
            new_sqref_str = " ".join(new_ranges)
            dv.sqref = new_sqref_str
            report.warnings.append(
                f"{ws.title}: expanded validation sqref {old_sqref} → {new_sqref_str}"
            )


# ─────────────────── x14 Extension restoration ───────────────────

# openpyxl strips <extLst> containing x14:dataValidations on load.
# After save, we patch the output .xlsx (a ZIP) to transplant x14 blocks
# from the original template, adjusting sqref ranges for actual data rows.

_X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_XM_NS = "http://schemas.microsoft.com/office/excel/2006/main"
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

# Pre-register namespaces so ET.tostring doesn't create ns0/ns1 prefixes
ET.register_namespace("", _SHEET_NS)
ET.register_namespace("x14", _X14_NS)
ET.register_namespace("xm", _XM_NS)


def _get_sheet_xml_map(template_path: Path) -> dict[str, str]:
    """Map sheet names to their xl/worksheets/sheet*.xml paths.

    Reads xl/workbook.xml to get sheet order + names, then maps to
    the rId → file target via xl/_rels/workbook.xml.rels.
    """
    ns_wb = {"x": _SHEET_NS}

    with zipfile.ZipFile(template_path, "r") as zf:
        # Parse workbook.xml for sheet names + rIds
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = wb_xml.findall(".//x:sheets/x:sheet", ns_wb)
        rid_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        name_to_rid = {}
        for s in sheets:
            name_to_rid[s.get("name")] = s.get(f"{{{rid_ns}}}id")

        # Parse rels for rId → target
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {}
        for rel in rels_xml:
            rid_to_target[rel.get("Id")] = "xl/" + rel.get("Target")

    return {name: rid_to_target[rid] for name, rid in name_to_rid.items() if rid in rid_to_target}


def _extract_x14_ext_block(xml_bytes: bytes) -> str | None:
    """Extract the full <ext> block containing x14:dataValidations from sheet XML."""
    root = ET.fromstring(xml_bytes)
    for ext_lst in root.findall(f"{{{_SHEET_NS}}}extLst"):
        for ext in ext_lst:
            # Check if this ext contains x14:dataValidations
            for child in ext:
                if f"{{{_X14_NS}}}dataValidation" in child.tag or "dataValidations" in child.tag:
                    return ET.tostring(ext, encoding="unicode")
    return None


def _adjust_x14_sqref(ext_xml: str, start_row: int, last_row: int) -> str:
    """Expand xm:sqref ranges in x14 extension to cover all data rows.

    Template may have fragmented ranges (e.g. H25:H27 H29:H31 H33:H34).
    We collapse them into a single contiguous range per column.
    """
    ext = ET.fromstring(ext_xml)

    for sqref_el in ext.iter(f"{{{_XM_NS}}}sqref"):
        if sqref_el.text:
            parts = sqref_el.text.strip().split()
            seen_cols: set[str] = set()
            new_parts = []
            for part in parts:
                m = re.match(r"^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", part)
                if m:
                    col_s = m.group(1)
                    col_e = m.group(3) or col_s
                    col_key = f"{col_s}:{col_e}"
                    if col_key in seen_cols:
                        continue  # skip duplicate column ranges
                    seen_cols.add(col_key)
                    new_parts.append(f"{col_s}{start_row}:{col_e}{last_row}")
                else:
                    new_parts.append(part)
            sqref_el.text = " ".join(new_parts)

    return ET.tostring(ext, encoding="unicode")


def _restore_x14_extensions(
    template_path: Path,
    output_path: Path,
    schema: dict,
    data: dict,
    report: FillReport,
):
    """Patch output xlsx to restore x14:dataValidations from template.

    After openpyxl saves (stripping extLst), we:
    1. Read each sheet's x14 ext block from the template
    2. Adjust sqref for actual data row count
    3. Inject into the corresponding sheet XML in the output ZIP
    """
    sheets_schema = schema.get("sheets") or {}
    patches: dict[str, str] = {}  # xml_path → adjusted ext block

    # Build sheet map + read sheet XMLs in a single template ZIP open
    ns_wb = {"x": _SHEET_NS}
    rid_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    with zipfile.ZipFile(template_path, "r") as zf:
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheet_nodes = wb_xml.findall(".//x:sheets/x:sheet", ns_wb)
        name_to_rid = {s.get("name"): s.get(f"{{{rid_ns}}}id") for s in sheet_nodes}
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {rel.get("Id"): "xl/" + rel.get("Target") for rel in rels_root}
        sheet_map = {
            name: rid_to_target[rid] for name, rid in name_to_rid.items() if rid in rid_to_target
        }
        for sheet_name, xml_path in sheet_map.items():
            if sheet_name not in sheets_schema:
                continue
            dt = sheets_schema[sheet_name].get("data_table")
            try:
                xml_bytes = zf.read(xml_path)
            except KeyError:
                continue
            ext_block = _extract_x14_ext_block(xml_bytes)
            if not ext_block:
                continue

            # Compute actual last data row
            if dt:
                start_row = dt["start_row"]
                rows_data = resolve(data, dt["source"]) or []
                last_row = start_row + len(rows_data) - 1 if rows_data else start_row
                ext_block = _adjust_x14_sqref(ext_block, start_row, last_row)

            patches[xml_path] = ext_block

    if not patches:
        return

    # Rebuild ZIP with patched sheet XMLs
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        with (
            zipfile.ZipFile(output_path, "r") as zf_in,
            zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zf_out,
        ):
            for item in zf_in.infolist():
                raw = zf_in.read(item.filename)
                if item.filename in patches:
                    raw = _inject_ext_block(raw, patches[item.filename])
                    report.warnings.append(
                        f"Restored x14 dataValidation extension in {item.filename}"
                    )
                zf_out.writestr(item, raw)
        shutil.move(tmp_path, output_path)
    finally:
        if Path(tmp_path).exists():
            Path(tmp_path).unlink()


def _inject_ext_block(sheet_xml_bytes: bytes, ext_block_str: str) -> bytes:
    """Inject an <ext> block into a sheet's <extLst>, creating it if absent."""
    # Parse, preserving declaration
    root = ET.fromstring(sheet_xml_bytes)

    # Find or create extLst
    ext_lst = root.find(f"{{{_SHEET_NS}}}extLst")
    if ext_lst is None:
        ext_lst = ET.SubElement(root, f"{{{_SHEET_NS}}}extLst")

    # Remove any existing x14 dataValidation ext blocks (shouldn't be any, but safe)
    for ext in list(ext_lst):
        for child in ext:
            if _X14_NS in child.tag:
                ext_lst.remove(ext)
                break

    # Append the new ext block
    new_ext = ET.fromstring(ext_block_str)
    ext_lst.append(new_ext)

    # Serialize back — must include XML declaration for xlsx
    out = ET.tostring(root, encoding="unicode", xml_declaration=False)
    return b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + out.encode("utf-8")


# ─────────────────────────── Main ───────────────────────────


def fill(template_path: Path, schema_path: Path, data_path: Path, output_path: Path) -> FillReport:
    report = FillReport()

    # Load
    schema = yaml.safe_load(schema_path.read_text(encoding="utf-8"))
    data = json.loads(data_path.read_text(encoding="utf-8"))

    # Copy template first so we never mutate original
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path, data_only=False, keep_vba=False, keep_links=True)

    # Pre-flight: required sheets
    required = schema.get("required_sheets", [])
    missing = [s for s in required if s not in wb.sheetnames]
    if missing:
        report.validator_failures.append(f"Missing required sheets: {missing}")
        return report

    # Process each declared sheet
    for sheet_name, sheet_schema in (schema.get("sheets") or {}).items():
        if sheet_name not in wb.sheetnames:
            report.warnings.append(f"Schema declares sheet '{sheet_name}' but not in workbook")
            continue
        ws = wb[sheet_name]
        process_sheet_writes(ws, sheet_schema, data, schema, report)
        if "data_table" in sheet_schema:
            process_data_table(ws, sheet_schema, data, schema, report)
            _expand_data_validations(ws, sheet_schema, data, report)

    wb.save(output_path)

    # Restore x14 extension data validations (stripped by openpyxl on load)
    _restore_x14_extensions(template_path, output_path, schema, data, report)

    # Validate (use in-memory workbook — avoids redundant reload from disk)
    run_validators(wb, schema, report)

    return report


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--schema", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--report", default=None)
    args = ap.parse_args()

    report = fill(
        Path(args.template),
        Path(args.schema),
        Path(args.data),
        Path(args.output),
    )

    report_dict = report.to_dict()
    if args.report:
        Path(args.report).parent.mkdir(parents=True, exist_ok=True)
        Path(args.report).write_text(
            json.dumps(report_dict, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    # Human-readable summary
    print(f"Fill: {report.writes_done} cells, {report.rows_written} data rows")
    if report.warnings:
        print(f"Warnings: {len(report.warnings)}")
        for w in report.warnings[:5]:
            print(f"  - {w}")
    if report.validator_failures:
        print(f"VALIDATION FAILURES: {len(report.validator_failures)}")
        for f in report.validator_failures:
            print(f"  ✗ {f}")
        sys.exit(1)
    print("OK")


if __name__ == "__main__":
    main()
